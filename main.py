from PIL import Image
import openpyxl, os
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

scale = 10  # px per block
cellSize = 4


def avgBlock(img, pix, x, y):
    res = [0, 0, 0]
    cnt = 0
    for w in range(x, min(img.size[0], x + scale)):
        for h in range(y, min(img.size[1], y + scale)):
            res = [x + y for x, y in zip(res, pix[w, h])]
            cnt += 1
    res = [x // cnt for x in res]
    return res


def IMGtoXLSX(img, ws):
    imgW, imgH = img.size
    pix = img.load()
    for w in range(0, imgW // scale):
        ws.column_dimensions[get_column_letter(w + 1)].width = cellSize / 5.25
        for h in range(0, imgH // scale):
            ws.row_dimensions[h + 1].height = cellSize
            avg = avgBlock(img, pix, w * scale, h * scale)
            avgColor = '%02x%02x%02x' % (avg[0], avg[1], avg[2])
            ws.cell(h + 1, w + 1).fill = PatternFill(fgColor=avgColor, fill_type="solid")


imgName = '6.jpg'
img = Image.open(imgName)
wb = openpyxl.Workbook()
ws = wb.get_active_sheet()
ws.title = 'IMAGE'
IMGtoXLSX(img, ws)
wb.save(os.path.splitext(imgName)[0] + '.xlsx')
